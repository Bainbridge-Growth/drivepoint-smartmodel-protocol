"""
SmartModel Root Template Generator
Produces a blank-slate .xlsx that is a valid SmartModel Protocol v6.0 workbook.

Strategy: Build Index + Settings sheets with openpyxl, then rebuild the zip
from scratch — copying the webextension structure byte-for-byte from a known-good
production SmartModel, fixing openpyxl quirks (absolute paths, missing XML decls),
and bundling the smartmodel/ directory.

Includes:
  - Settings tab (dark gray)
  - Index tab (white)
  - smartmodel/skills/ and smartmodel/imports/ bundled in zip
  - WebExtension XML to prompt Drivepoint add-in install on open
    (Microsoft AppSource: WA200007015)

Usage:
    python3 tools/generate_root_template.py
Output:
    templates/smartmodel_root_v1.0.0.xlsx
"""

import os
import re
import zipfile
import uuid
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ADDIN_ASSET_ID   = "WA200007015"   # Microsoft AppSource product ID
ADDIN_STORE      = "en-US"
ADDIN_STORE_TYPE = "OMEX"          # Office Marketplace

PROTOCOL_VERSION = "6.0"
MODEL_VERSION    = "1.0.0"
MODEL_NAME       = "SmartModel Root Template"
SKILL_FILE_PATH  = "smartmodel/skills/root-template-skill.md"
IMPORTS_FILE_PATH = "smartmodel/imports/imports.yaml"

OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "templates", "smartmodel_root_v1.0.0.xlsx"
)

# ---------------------------------------------------------------------------
# Colors / styles (per Template Rules v3.2)
# ---------------------------------------------------------------------------

BLUE_TAB      = "64B1FF"   # Row 1 / title bar
BLACK         = "000000"
WHITE         = "FFFFFF"
GRAY_MID      = "808080"   # Row 3
SECTION_BLUE  = "63AEFF"   # Section border
BLUE_BULLET   = "4472C4"

# Tab colors
TAB_DARK_GRAY  = "404040"  # Settings
TAB_WHITE      = "FFFFFF"  # Index

fill_blue   = PatternFill(start_color=BLUE_TAB, end_color=BLUE_TAB, fill_type="solid")
fill_black  = PatternFill(start_color=BLACK,    end_color=BLACK,    fill_type="solid")
fill_gray   = PatternFill(start_color=GRAY_MID, end_color=GRAY_MID, fill_type="solid")
fill_light  = PatternFill(patternType="solid",
                          fgColor=Color(theme=0, tint=-0.0499893185216834))
fill_settings_row = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

font_white_14  = Font(name="Calibri", size=14, color=WHITE, bold=False)
font_white_11b = Font(name="Calibri", size=11, color=WHITE, bold=True)
font_white_11  = Font(name="Calibri", size=11, color=WHITE, bold=False)
font_black_14b = Font(name="Calibri", size=14, color=BLACK, bold=True)
font_black_11  = Font(name="Calibri", size=11, color=BLACK, bold=False)
font_black_11b = Font(name="Calibri", size=11, color=BLACK, bold=True)
font_gray_11i  = Font(name="Calibri", size=11, color=GRAY_MID, italic=True)
font_mono      = Font(name="Menlo",   size=10, color=BLACK)
font_bullet_blue = Font(name="Calibri", size=11, color=BLUE_BULLET)

border_section = Border(bottom=Side(style="thick", color=SECTION_BLUE))

right_align = Alignment(horizontal="right")


def section_border(ws, row, last_col):
    for col in range(2, last_col + 1):
        ws.cell(row, col).border = border_section


# ---------------------------------------------------------------------------
# Settings tab
# ---------------------------------------------------------------------------

SETTINGS_ROWS = [
    # (id, setting, value, description)
    ("settings.smartmodelSpec",      "Protocol Version",     PROTOCOL_VERSION,  "SmartModel Protocol version"),
    ("settings.modelVersion",        "Model Version",        MODEL_VERSION,      "Template version (semver)"),
    ("settings.modelName",           "Model Name",           MODEL_NAME,         "Human-readable model name"),
    ("settings.modelType",           "Model Type",           "template",         '"template" or "model"'),
    ("settings.modelStartDate",      "ProForma Start Date",  "",                 "Start date for forecast period"),
    ("settings.historicalStartDate", "Historical Start",     "",                 "Start date for historical data"),
    ("settings.companyId",           "Company ID",           "",                 "Drivepoint company identifier"),
    ("settings.companyName",         "Company Name",         "",                 "Company name"),
    ("settings.currency",            "Currency",             "USD",              "Base currency (ISO 4217)"),
    ("settings.author",              "Author",               "",                 "Author name"),
    ("settings.authorId",            "Author ID",            "",                 "Author identifier"),
    ("settings.skillFile",           "Skill File",           SKILL_FILE_PATH,    "Path to template skill .md"),
    ("settings.importsFile",         "Imports File",         IMPORTS_FILE_PATH,  "Path to imports.yaml"),
]


def build_settings_tab(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = TAB_DARK_GRAY

    # Column widths
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 44

    # Header row
    headers = ["id", "setting", "value", "description"]
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.fill = header_fill
        cell.font = Font(name="Menlo", size=10, color=WHITE, bold=True)

    # Data rows
    for row_idx, (sid, setting, value, description) in enumerate(SETTINGS_ROWS, 2):
        row_fill = fill_light if row_idx % 2 == 0 else PatternFill()

        id_cell = ws.cell(row_idx, 1, sid)
        id_cell.font = font_mono
        id_cell.fill = row_fill

        setting_cell = ws.cell(row_idx, 2, setting)
        setting_cell.font = font_black_11
        setting_cell.fill = row_fill

        val_cell = ws.cell(row_idx, 3, value)
        val_cell.font = font_black_11
        val_cell.fill = row_fill

        desc_cell = ws.cell(row_idx, 4, description)
        desc_cell.font = font_gray_11i
        desc_cell.fill = row_fill


# ---------------------------------------------------------------------------
# Index tab
# ---------------------------------------------------------------------------

def build_index_tab(wb):
    ws = wb.active
    ws.title = "Index"
    ws.sheet_properties.tabColor = TAB_WHITE

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 48

    # Row 1 — blue header bar
    for col in range(1, 20):
        ws.cell(1, col).fill = fill_blue
        ws.cell(1, col).font = font_white_14
    ws["A1"] = "≡"
    ws["C1"] = MODEL_NAME
    ws["C1"].font = font_white_14

    # Row 2 — black bar (date spine placeholder)
    for col in range(1, 20):
        ws.cell(2, col).fill = fill_black
        ws.cell(2, col).font = font_white_11b
    ws["C2"] = "SmartModel Index"

    # Row 3 — gray bar
    for col in range(1, 20):
        ws.cell(3, col).fill = fill_gray
        ws.cell(3, col).font = font_white_11

    # Row 4 — status bar
    for col in range(1, 20):
        ws.cell(4, col).fill = fill_light
    ws["A4"] = "•"
    ws["A4"].font = font_bullet_blue
    ws["C4"] = "Add-in managed — do not edit manually"
    ws["C4"].font = font_black_11

    # Rows 5-6 blank

    # Row 7 — title
    ws["C7"] = MODEL_NAME
    ws["C7"].font = font_black_14b
    section_border(ws, 7, 10)

    # Row 8 — description
    ws["C8"] = "Table of contents — maintained by the Drivepoint add-in"
    ws["C8"].font = font_gray_11i

    # Row 9 — metadata anchor
    ws["B9"] = "metadata___name"
    ws["B9"].font = font_mono
    ws["C9"] = "Name"
    ws["C9"].font = font_black_11
    ws["D9"] = MODEL_NAME
    ws["D9"].font = font_black_11

    # Update C1/C7 to reference D9
    ws["C1"] = "=D9"
    ws["C7"] = "=D9"

    # Placeholder section
    ws["C11"] = "Templates"
    ws["C11"].font = font_black_11b
    section_border(ws, 11, 10)
    ws["C12"] = "No templates registered. Open a SmartModel schedule sheet to register a template."
    ws["C12"].font = font_gray_11i


# ---------------------------------------------------------------------------
# Bundled smartmodel/ file content
# ---------------------------------------------------------------------------

PLACEHOLDER_SKILL = """\
---
name: root-template
description: Root SmartModel template — blank slate with minimum required structure. Replace this file with a template-specific skill.
user-invocable: false
---

# Root Template Skill

This is a placeholder skill for the SmartModel root template. Replace this file with a
template-specific skill that teaches the agent how to operate the model.

See the SmartModel Protocol skill for universal grammar:
https://raw.githubusercontent.com/Bainbridge-Growth/drivepoint-smartmodel-protocol/main/protocol/v6.0/smartmodel-skill.md
"""

PLACEHOLDER_IMPORTS = """\
# SmartModel imports declaration
# See imports schema: https://github.com/Bainbridge-Growth/drivepoint-smartmodel-protocol
imports: []
"""


# ---------------------------------------------------------------------------
# WebExtension XML (Drivepoint add-in, AppSource WA200007015)
# Correct structure reverse-engineered from production SmartModel files.
#
# File layout inside xlsx zip:
#   _rels/.rels                                → references taskpanes.xml
#   xl/webextensions/taskpanes.xml             → lists taskpanes
#   xl/webextensions/_rels/taskpanes.xml.rels  → references webextension1.xml
#   xl/webextensions/webextension1.xml         → add-in manifest reference
# ---------------------------------------------------------------------------

WE_GUID = str(uuid.uuid4()).upper()
ADDIN_VERSION = "6.0.16.0"

WEBEXTENSION_XML = f"""\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<we:webextension xmlns:we="http://schemas.microsoft.com/office/webextensions/webextension/2010/11" id="{{{WE_GUID}}}"><we:reference id="{ADDIN_ASSET_ID.lower()}" version="{ADDIN_VERSION}" store="{ADDIN_STORE}" storeType="{ADDIN_STORE_TYPE}"/><we:alternateReferences><we:reference id="{ADDIN_ASSET_ID.lower()}" version="{ADDIN_VERSION}" store="" storeType="{ADDIN_STORE_TYPE}"/></we:alternateReferences><we:properties/><we:bindings/><we:snapshot xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/><we:extLst><a:ext xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" uri="{{D87F86FE-615C-45B5-9D79-34F1136793EB}}"><we:containsCustomFunctions/></a:ext><a:ext xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" uri="{{0858819E-0033-43BF-8937-05EC82904868}}"><we:backgroundApp state="1" runtimeId="Taskpane.Url"/></a:ext></we:extLst></we:webextension>"""

# namespace: wetp (not tp) — matches production files
TASKPANE_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<wetp:taskpanes xmlns:wetp="http://schemas.microsoft.com/office/webextensions/taskpanes/2010/11"><wetp:taskpane dockstate="right" visibility="0" width="350" row="0"><wetp:webextensionref xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1"/></wetp:taskpane></wetp:taskpanes>"""

TASKPANE_RELS_XML = """\
<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.microsoft.com/office/2011/relationships/webextension" Target="webextension1.xml"/></Relationships>"""

# Injected into _rels/.rels (root), NOT xl/_rels/workbook.xml.rels
ROOT_TASKPANE_REL = (
    '<Relationship Id="rId_dp1" '
    'Type="http://schemas.microsoft.com/office/2011/relationships/webextensiontaskpanes" '
    'Target="xl/webextensions/taskpanes.xml"/>'
)

# Content type overrides → correct paths under /xl/webextensions/
CONTENT_TYPE_WEBEXT = (
    '<Override PartName="/xl/webextensions/webextension1.xml" '
    'ContentType="application/vnd.ms-office.webextension+xml"/>'
)
CONTENT_TYPE_TASKPANE = (
    '<Override PartName="/xl/webextensions/taskpanes.xml" '
    'ContentType="application/vnd.ms-office.webextensiontaskpanes+xml"/>'
)


# ---------------------------------------------------------------------------
# Post-process: inject web extension + smartmodel/ bundle into xlsx zip
# ---------------------------------------------------------------------------

def fix_workbook_rels(data: bytes) -> bytes:
    """Fix openpyxl absolute paths (/xl/worksheets/...) → relative (worksheets/...)."""
    text = data.decode("utf-8")
    text = text.replace('Target="/xl/', 'Target="')
    return text.encode("utf-8")


def ensure_xml_decl(data: bytes) -> bytes:
    """Add XML declaration if missing (openpyxl omits it)."""
    decl = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
    if not data.lstrip().startswith(b"<?xml"):
        return decl + data
    return data


def inject_into_xlsx(src_path: str, dst_path: str):
    """
    Rebuilds xlsx from openpyxl output:
      - Fixes absolute paths in workbook.xml.rels
      - Adds XML declarations to all XML parts
      - Injects webextension into _rels/.rels and [Content_Types].xml
      - Bundles smartmodel/ directory
    """
    buf = BytesIO()

    with zipfile.ZipFile(src_path, "r") as zin, \
         zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:

        for item in zin.infolist():
            data = zin.read(item.filename)
            name = item.filename

            if name.endswith(".xml") or name.endswith(".rels"):
                # Fix workbook.xml.rels absolute paths
                if name == "xl/_rels/workbook.xml.rels":
                    data = fix_workbook_rels(data)

                # Inject taskpane rel into _rels/.rels
                if name == "_rels/.rels":
                    text = data.decode("utf-8")
                    text = text.replace(
                        "</Relationships>",
                        f"{ROOT_TASKPANE_REL}</Relationships>"
                    )
                    data = text.encode("utf-8")

                # Inject content type overrides
                if name == "[Content_Types].xml":
                    text = data.decode("utf-8")
                    text = text.replace(
                        "</Types>",
                        f"{CONTENT_TYPE_WEBEXT}{CONTENT_TYPE_TASKPANE}</Types>"
                    )
                    data = text.encode("utf-8")

                data = ensure_xml_decl(data)

            zout.writestr(item, data)

        # Add webextension files (all under xl/webextensions/)
        zout.writestr("xl/webextensions/webextension1.xml",        WEBEXTENSION_XML)
        zout.writestr("xl/webextensions/taskpanes.xml",            TASKPANE_XML)
        zout.writestr("xl/webextensions/_rels/taskpanes.xml.rels", TASKPANE_RELS_XML)

        # smartmodel bundle
        zout.writestr(SKILL_FILE_PATH,   PLACEHOLDER_SKILL)
        zout.writestr(IMPORTS_FILE_PATH, PLACEHOLDER_IMPORTS)

    # Write final file
    os.makedirs(os.path.dirname(dst_path), exist_ok=True)
    with open(dst_path, "wb") as f:
        f.write(buf.getvalue())


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = openpyxl.Workbook()

    build_index_tab(wb)
    build_settings_tab(wb)

    # Save to temp file first (openpyxl needs a real path)
    tmp_path = OUTPUT_PATH + ".tmp"
    os.makedirs(os.path.dirname(tmp_path), exist_ok=True)
    wb.save(tmp_path)

    # Inject web extension + smartmodel bundle
    inject_into_xlsx(tmp_path, OUTPUT_PATH)
    os.remove(tmp_path)

    print(f"✓ Generated: {OUTPUT_PATH}")

    # Verify bundle
    with zipfile.ZipFile(OUTPUT_PATH, "r") as z:
        names = z.namelist()
        checks = [
            "xl/webextensions/webextension1.xml",
            "xl/webextensions/taskpanes.xml",
            "xl/webextensions/_rels/taskpanes.xml.rels",
            SKILL_FILE_PATH,
            IMPORTS_FILE_PATH,
        ]
        print("\nBundle verification:")
        for c in checks:
            status = "✓" if c in names else "✗ MISSING"
            print(f"  {status}  {c}")


if __name__ == "__main__":
    main()
